from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = [
    "Order ID",
    "Tanggal Order",
    "Perusahaan",
    "Alamat Jemput",
    "Alamat Antar",
    "Jenis Barang",
    "Berat (kg)",
    "Tipe Kendaraan",
    "Driver",
    "Plat Nomor",
    "Status",
    "Total Harga",
    "Payout Driver",
    "Komisi Platform",
    "Tgl Selesai",
]


def build_orders_report(orders: list[dict]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Order"

    ws.append(HEADERS)
    header_fill = PatternFill(start_color="435058", end_color="435058", fill_type="solid")
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for o in orders:
        ws.append(
            [
                o["id"],
                o["created_at"],
                o["company_name"],
                o["pickup_address"],
                o["dropoff_address"],
                o["cargo_type"],
                float(o["weight_kg"]),
                o["vehicle_type_requested"],
                o["driver_name"],
                o["plate_number"],
                o["status"],
                float(o["total_price"]),
                float(o["driver_payout"]),
                float(o["platform_commission"]),
                o["delivered_at"],
            ]
        )

    for col_idx in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
